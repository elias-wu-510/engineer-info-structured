#!/usr/bin/env python3
"""Generate a simple 2-row Excel for Daily Report labour codes only.

Output format:
  Row 1: trade names for code 1-34 and B1-B12
  Row 2: access-gate headcount mapped to each code

Non-mapped/new trades are excluded by design.
"""
from __future__ import annotations
import argparse, json, re
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def norm(v):
    return re.sub(r"\s+", "", str(v or "")).strip()

def to_int(v):
    if v is None: return 0
    if isinstance(v, (int,float)): return int(v)
    m = re.search(r"-?\d+", str(v).replace(',', ''))
    return int(m.group()) if m else 0

def strip_trade_prefix(v):
    """Remove leading access-report job code prefixes before matching staff titles.

    Examples: ADM060-助理地盤行政經理, PM-030-結構工程師, SE-003-安全經理.
    """
    text = str(v or "").strip()
    return re.sub(r"^[A-Za-z]+-?\d+[-－–—]\s*", "", text).strip()

def load_cic_mapping(wb):
    ws = wb['CIC工種對應表']
    by_company_trade = {}
    by_trade_candidates = defaultdict(list)
    for r in range(2, ws.max_row+1):
        code = ws.cell(r,1).value
        cic = ws.cell(r,2).value
        co = ws.cell(r,3).value
        trade = ws.cell(r,4).value
        if code is None or not trade:
            continue
        item = {'daily_code': str(code).strip().upper(), 'cic_trade': cic, 'company': co, 'trade': trade, 'source_row': r}
        if co:
            by_company_trade[norm(co) + norm(trade)] = item
        by_trade_candidates[norm(trade)].append(item)
    by_trade = {}
    ambiguous_by_trade = {}
    for trade_key, items in by_trade_candidates.items():
        codes = {i['daily_code'] for i in items}
        if len(codes) == 1:
            by_trade[trade_key] = items[0]
        else:
            ambiguous_by_trade[trade_key] = items
    return by_company_trade, by_trade, ambiguous_by_trade


def load_china_state_staff_mapping(wb):
    """Load 中建工種 sheet: trade title -> Staff S code via abbreviation.

    The sheet columns are 工種, 職位名稱 where 職位名稱 is an abbreviation
    such as PM/AM/E/AE/BS/QS/Safety/etc.  These map to S1-S16.
    """
    sheet_name = '中建工種' if '中建工種' in wb.sheetnames else None
    if not sheet_name:
        return {}
    abbr_to_s = {
        'PM':'S1', 'AM':'S2', 'SITEAGENT':'S3', 'GENERAL':'S4',
        'FOREMAN':'S5', 'E':'S6', 'AE':'S6', 'BS':'S7', 'PLANNING':'S8',
        'QUALITY':'S9', 'BIM':'S10', 'QS':'S11', 'SAFETY':'S12',
        'ENVI':'S13', 'LABOUR':'S14', 'WATCHMAN':'S15',
        'OTHER':'S16',
    }
    ws = wb[sheet_name]
    mapping = {}
    for r in range(2, ws.max_row+1):
        trade = ws.cell(r,1).value
        abbr = ws.cell(r,2).value
        if not trade or not abbr:
            continue
        code = abbr_to_s.get(norm(abbr).upper())
        if code:
            mapping[norm(trade)] = {'daily_code': code, 'source': sheet_name, 'source_row': r}
    return mapping

def load_daily_code_names(wb):
    ws = wb['daily report碼表']
    names = {}
    for r in range(2, ws.max_row+1):
        if ws.cell(r,1).value:
            names[str(ws.cell(r,1).value).strip().upper()] = ws.cell(r,2).value
        if ws.cell(r,3).value:
            names[str(ws.cell(r,3).value).strip().upper()] = ws.cell(r,4).value
        if ws.cell(r,5).value:
            names[str(ws.cell(r,5).value).strip().upper()] = ws.cell(r,6).value
    return names

def wanted_codes():
    return [str(i) for i in range(1,35)] + [f'B{i}' for i in range(1,13)] + [f'S{i}' for i in range(1,17)]

def parse_access_report(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records=[]; company=None; report_date=ws.cell(5,2).value
    for r in range(6, ws.max_row+1):
        name = ws.cell(r,1).value
        if name is None: continue
        name_s = str(name).strip()
        if not name_s: continue
        if name_s == '累計':
            company = None; continue
        count = to_int(ws.cell(r,2).value)
        if count == 0 and ws.cell(r,2).value in (None, ''):
            company = name_s; continue
        if company and count:
            records.append({'company': company, 'trade': name_s, 'count': count, 'row': r})
    return records, report_date

def generate(mapping_path, access_path, output_path, summary_path=None):
    map_wb = load_workbook(mapping_path, data_only=False)
    cic_map, cic_by_trade, ambiguous_by_trade = load_cic_mapping(map_wb)
    china_state_staff_by_trade = load_china_state_staff_mapping(map_wb)
    code_names = load_daily_code_names(map_wb)
    records, report_date = parse_access_report(access_path)
    allowed = set(wanted_codes())
    counts = Counter(); excluded=[]; matched=[]
    for rec in records:
        key = norm(rec['company']) + norm(rec['trade'])
        m = cic_map.get(key)
        method = 'company_trade'
        if not m:
            tk = norm(rec['trade'])
            if tk in cic_by_trade:
                m = cic_by_trade[tk]; method = 'trade_fallback'
            elif tk in china_state_staff_by_trade:
                m = china_state_staff_by_trade[tk]; method = 'china_state_staff_trade'
            elif norm(strip_trade_prefix(rec['trade'])) in china_state_staff_by_trade:
                m = china_state_staff_by_trade[norm(strip_trade_prefix(rec['trade']))]; method = 'china_state_staff_trade_strip_prefix'
            elif tk in ambiguous_by_trade:
                excluded.append({**rec, 'reason':'ambiguous_or_not_in_1-34_B1-B12_S1-S16'})
                continue
            else:
                excluded.append({**rec, 'reason':'no_mapping_or_not_in_1-34_B1-B12_S1-S16'})
                continue
        code = str(m['daily_code']).strip().upper()
        if code not in allowed:
            excluded.append({**rec, 'reason':'code_not_required', 'daily_code': code})
            continue
        counts[code] += rec['count']
        matched.append({**rec, 'daily_code': code, 'method': method})

    wb = Workbook(); ws = wb.active; ws.title = 'Daily Report Codes'
    thin = Side(style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_block(codes, name_row, count_row, fill, total_label):
        total = sum(counts.get(code, 0) for code in codes)
        ws.cell(name_row, 1).value = total_label
        ws.cell(count_row, 1).value = total
        for c in (ws.cell(name_row, 1), ws.cell(count_row, 1)):
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(name_row, 1).font = Font(bold=True)
        ws.cell(name_row, 1).fill = PatternFill('solid', fgColor='FFD966')
        ws.column_dimensions['A'].width = 16

        for idx, code in enumerate(codes, 2):
            c1=ws.cell(name_row,idx); c2=ws.cell(count_row,idx)
            c1.value = code_names.get(code) or code
            c2.value = counts.get(code, 0)
            for c in (c1,c2):
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c1.font = Font(bold=True)
            c1.fill = PatternFill('solid', fgColor=fill)
            ws.column_dimensions[c1.column_letter].width = 18

    labour_codes = [str(i) for i in range(1,35)]
    bs_codes = [f'B{i}' for i in range(1,13)]
    staff_codes = [f'S{i}' for i in range(1,17)]
    write_block(labour_codes, 1, 2, 'D9EAF7', 'Labour total人數')
    write_block(bs_codes, 3, 4, 'E2F0D9', 'BS Labour total人數')
    write_block(staff_codes, 5, 6, 'FCE4D6', 'Staff total人數')
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[3].height = 60
    ws.row_dimensions[5].height = 60
    ws.freeze_panes = 'A2'
    output_path = Path(output_path); output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    summary = {
        'report_date_header': report_date,
        'access_total': sum(r['count'] for r in records),
        'included_total': sum(counts.values()),
        'excluded_total': sum(r['count'] for r in excluded),
        'matched_record_count': len(matched),
        'excluded_record_count': len(excluded),
        'counts_by_code': dict(counts),
        'excluded': excluded,
    }
    if summary_path:
        Path(summary_path).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    return summary

if __name__ == '__main__':
    ap=argparse.ArgumentParser()
    ap.add_argument('--mapping', required=True)
    ap.add_argument('--access-report', required=True)
    ap.add_argument('--output', required=True)
    ap.add_argument('--summary')
    args=ap.parse_args()
    s=generate(args.mapping,args.access_report,args.output,args.summary)
    print(json.dumps({k:s[k] for k in ['access_total','included_total','excluded_total','matched_record_count','excluded_record_count']}, ensure_ascii=False, indent=2))
