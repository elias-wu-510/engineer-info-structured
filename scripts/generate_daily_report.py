#!/usr/bin/env python3
"""Generate Daily Report Excel from access gate headcount report.

Usage:
  python scripts/generate_daily_report.py --mapping mapping.xlsx --access-report gate.xlsx --output out.xlsx
"""
from __future__ import annotations
import argparse, json, re
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path
from openpyxl import load_workbook

LABOUR_START_COL = 24  # X, codes 1-34 -> X:BE
BS_START_COL = 75      # BW, codes B1-B12 -> BW:CH
STAFF_START_COL = 2    # B, codes S1-S16 -> B:Q
LABOUR_ROW = 19        # blank Site Activities input row in demo
STAFF_ROW = 46         # Personnel on site today input row in demo


def norm(v):
    return re.sub(r"\s+", "", str(v or "")).strip()

def to_int(v):
    if v is None: return 0
    if isinstance(v, (int,float)): return int(v)
    m = re.search(r"-?\d+", str(v).replace(',', ''))
    return int(m.group()) if m else 0

def copy_cell_style(src, dst):
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format: dst.number_format = src.number_format
    if src.font: dst.font = copy(src.font)
    if src.fill: dst.fill = copy(src.fill)
    if src.border: dst.border = copy(src.border)
    if src.alignment: dst.alignment = copy(src.alignment)
    if src.protection: dst.protection = copy(src.protection)


def load_cic_mapping(wb):
    ws = wb['CIC工種對應表']
    by_company_trade = {}
    by_trade_candidates = defaultdict(list)
    for r in range(2, ws.max_row+1):
        code = ws.cell(r,1).value
        cic = ws.cell(r,2).value
        co = ws.cell(r,3).value
        trade = ws.cell(r,4).value
        if code is None or not trade:
            continue
        item = {'daily_code': code, 'cic_trade': cic, 'company': co, 'trade': trade, 'source_row': r}
        if co:
            by_company_trade[norm(co) + norm(trade)] = item
        by_trade_candidates[norm(trade)].append(item)

    # Only use trade-only fallback when all rows for the same trade resolve to one Daily Report code.
    by_trade = {}
    ambiguous_by_trade = {}
    for trade_key, items in by_trade_candidates.items():
        codes = {str(i['daily_code']).strip().upper() for i in items}
        if len(codes) == 1:
            by_trade[trade_key] = items[0]
        else:
            ambiguous_by_trade[trade_key] = items
    return by_company_trade, by_trade, ambiguous_by_trade

def load_daily_codes(wb):
    ws = wb['daily report碼表']
    codes = {}
    for r in range(2, ws.max_row+1):
        if ws.cell(r,1).value:
            codes[str(ws.cell(r,1).value).strip().upper()] = {'kind':'staff','desc':ws.cell(r,2).value}
        if ws.cell(r,3).value:
            codes[str(ws.cell(r,3).value).strip().upper()] = {'kind':'labour','desc':ws.cell(r,4).value}
        if ws.cell(r,5).value:
            codes[str(ws.cell(r,5).value).strip().upper()] = {'kind':'bs','desc':ws.cell(r,6).value}
    return codes

def parse_access_report(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records=[]; company=None; report_date=None
    # date from B5 header usually "18\nMo"
    report_date = ws.cell(5,2).value
    for r in range(6, ws.max_row+1):
        name = ws.cell(r,1).value
        if name is None: continue
        name_s = str(name).strip()
        if not name_s: continue
        if name_s == '累計':
            company = None; continue
        count = to_int(ws.cell(r,2).value)
        # Company rows have blank/no count and following trade rows have counts.
        if count == 0 and ws.cell(r,2).value in (None, ''):
            company = name_s; continue
        if company and count:
            records.append({'company': company, 'trade': name_s, 'count': count, 'row': r})
    return records, report_date

def code_to_target(code):
    s = str(code).strip().upper()
    if re.fullmatch(r'S\d+', s):
        n=int(s[1:]); return STAFF_ROW, STAFF_START_COL+n-1, 'staff'
    if re.fullmatch(r'B\d+', s):
        n=int(s[1:]); return LABOUR_ROW, BS_START_COL+n-1, 'bs'
    if re.fullmatch(r'\d+', s):
        n=int(s); return LABOUR_ROW, LABOUR_START_COL+n-1, 'labour'
    return None

def generate(mapping_path, access_path, output_path, summary_path=None):
    map_wb = load_workbook(mapping_path, data_only=False)
    cic_map, cic_by_trade, ambiguous_by_trade = load_cic_mapping(map_wb)
    daily_codes = load_daily_codes(map_wb)
    records, report_date = parse_access_report(access_path)

    counts = Counter(); unmatched=[]; matched=[]; fallback_matched=[]; ambiguous=[]
    for rec in records:
        key = norm(rec['company']) + norm(rec['trade'])
        m = cic_map.get(key)
        match_method = 'company_trade'
        if not m:
            trade_key = norm(rec['trade'])
            if trade_key in cic_by_trade:
                m = cic_by_trade[trade_key]
                match_method = 'trade_fallback'
            elif trade_key in ambiguous_by_trade:
                cand = ambiguous_by_trade[trade_key]
                ambiguous.append({**rec, 'candidate_codes': sorted({str(i['daily_code']).strip().upper() for i in cand})})
                unmatched.append({**rec, 'reason': 'ambiguous_trade_fallback'})
                continue
            else:
                unmatched.append({**rec, 'reason': 'no_mapping'})
                continue
        code = str(m['daily_code']).strip().upper()
        counts[code] += rec['count']
        row = {**rec, 'daily_code': code, 'cic_trade': m.get('cic_trade'), 'match_method': match_method}
        matched.append(row)
        if match_method == 'trade_fallback':
            fallback_matched.append(row)

    out_wb = load_workbook(mapping_path, data_only=False)
    ws = out_wb['daily report demo']
    ws.title = 'Daily Report'

    # Clear target input rows but preserve styles/formulas elsewhere.
    for c in range(LABOUR_START_COL, LABOUR_START_COL+34): ws.cell(LABOUR_ROW,c).value = None
    for c in range(BS_START_COL, BS_START_COL+12): ws.cell(LABOUR_ROW,c).value = None
    for c in range(STAFF_START_COL, STAFF_START_COL+16): ws.cell(STAFF_ROW,c).value = 0
    ws.cell(LABOUR_ROW, 2).value = '=B18+0.01'
    ws.cell(LABOUR_ROW, 3).value = 'Auto generated from access gate report'
    ws.cell(LABOUR_ROW, 23).value = 'M/S:'
    for c in range(1, ws.max_column+1):
        copy_cell_style(ws.cell(18,c), ws.cell(LABOUR_ROW,c))
    ws.cell(LABOUR_ROW, 89).value = f'=SUM(X{LABOUR_ROW}:CJ{LABOUR_ROW})'

    placed = {}; unsupported=[]
    for code, total in counts.items():
        target = code_to_target(code)
        if not target:
            unsupported.append({'code': code, 'count': total}); continue
        row, col, kind = target
        ws.cell(row, col).value = total
        placed[code] = {'count': total, 'cell': ws.cell(row,col).coordinate, 'kind': kind, 'desc': daily_codes.get(code,{}).get('desc')}

    # make sure formulas include row 19 for labour/BS totals (demo already does for X:BD; add BS total row formulas if blank)
    for c in range(LABOUR_START_COL, LABOUR_START_COL+34):
        ws.cell(20,c).value = f'=SUM({ws.cell(17,c).coordinate}:{ws.cell(19,c).coordinate})'
    for c in range(BS_START_COL, BS_START_COL+12):
        if ws.cell(20,c).value is None:
            ws.cell(20,c).value = f'=SUM({ws.cell(17,c).coordinate}:{ws.cell(19,c).coordinate})'

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)

    summary = {
        'report_date_header': report_date,
        'access_record_count': len(records),
        'matched_record_count': len(matched),
        'company_trade_matched_record_count': len(matched) - len(fallback_matched),
        'trade_fallback_matched_record_count': len(fallback_matched),
        'unmatched_record_count': len(unmatched),
        'ambiguous_record_count': len(ambiguous),
        'access_total': sum(r['count'] for r in records),
        'matched_total': sum(r['count'] for r in matched),
        'unmatched_total': sum(r['count'] for r in unmatched),
        'placed_total': sum(v['count'] for v in placed.values()),
        'placed_by_code': placed,
        'unsupported_codes': unsupported,
        'trade_fallback_matched': fallback_matched,
        'ambiguous': ambiguous,
        'unmatched': unmatched,
    }
    if summary_path:
        Path(summary_path).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    return summary

if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--mapping', required=True)
    ap.add_argument('--access-report', required=True)
    ap.add_argument('--output', required=True)
    ap.add_argument('--summary')
    ap.add_argument('--unmatched-csv', help='Optional CSV path for unmatched records')
    args=ap.parse_args()
    s=generate(args.mapping,args.access_report,args.output,args.summary)
    if args.unmatched_csv:
        import csv
        with open(args.unmatched_csv, 'w', newline='', encoding='utf-8-sig') as f:
            w=csv.DictWriter(f, fieldnames=['reason','company','trade','count','row','candidate_codes'])
            w.writeheader()
            for r in s.get('unmatched', []):
                rr={k:r.get(k,'') for k in ['reason','company','trade','count','row']}
                rr['candidate_codes']=';'.join(r.get('candidate_codes', [])) if isinstance(r.get('candidate_codes'), list) else ''
                w.writerow(rr)
    print(json.dumps({k:s[k] for k in ['access_record_count','matched_record_count','company_trade_matched_record_count','trade_fallback_matched_record_count','unmatched_record_count','ambiguous_record_count','access_total','matched_total','unmatched_total','placed_total']}, ensure_ascii=False, indent=2))
